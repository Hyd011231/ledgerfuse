"""导出：CSV（UTF-8-BOM）与 Excel（openpyxl 多 sheet 月报）。"""
from __future__ import annotations

import csv
import io

from ..db import get_conn
from . import stats

COLUMNS = ["id", "trans_time", "amount_yuan", "direction", "flow_type", "dup_status",
           "category", "counterparty", "description", "source", "account", "pay_method_raw",
           "trans_type_raw", "status_raw", "remark"]


def _query_rows(filters: dict) -> list[dict]:
    conn = get_conn()
    cond = ["t.is_deleted=0"]
    args: list = []
    if filters.get("date_from"):
        cond.append("t.trans_time>=?")
        args.append(filters["date_from"])
    if filters.get("date_to"):
        cond.append("t.trans_time<=?")
        args.append(filters["date_to"] + " 23:59:59")
    if filters.get("month"):
        cond.append("substr(t.trans_time,1,7)=?")
        args.append(filters["month"])
    if filters.get("activity_id"):
        cond.append("t.activity_id=?")
        args.append(filters["activity_id"])
    if filters.get("counted_only"):
        cond.append("t.status_ok=1 AND t.flow_type='normal' AND t.dup_status IN ('none','not_dup')")
    rows = conn.execute(f"""
        SELECT t.id, t.trans_time, t.amount/100.0 AS amount_yuan, t.direction, t.flow_type,
               t.dup_status, COALESCE(p.name||'/'||c.name, c.name, '') AS category,
               t.counterparty, t.description, t.source, COALESCE(a.name,'') AS account,
               t.pay_method_raw, t.trans_type_raw, t.status_raw, t.remark
        FROM transactions t
        LEFT JOIN categories c ON c.id=t.category_id
        LEFT JOIN categories p ON p.id=c.parent_id
        LEFT JOIN accounts a ON a.id=t.account_id
        WHERE {' AND '.join(cond)} ORDER BY t.trans_time DESC""", args).fetchall()
    return [dict(r) for r in rows]


def to_csv(filters: dict) -> bytes:
    rows = _query_rows(filters)
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=COLUMNS)
    w.writeheader()
    w.writerows(rows)
    return ("﻿" + buf.getvalue()).encode("utf-8")


def to_xlsx(filters: dict) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "交易明细"
    rows = _query_rows(filters)
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r[c] for c in COLUMNS])

    month = filters.get("month")
    ws2 = wb.create_sheet("分类汇总")
    ws2.append(["分类", "笔数", "金额(元)"])
    for r in stats.category_breakdown(month):
        ws2.append([r["name"], r["n"], round(r["total"] / 100, 2)])
    ws3 = wb.create_sheet("月度趋势")
    ws3.append(["月份", "收入(元)", "支出(元)", "结余(元)"])
    for r in stats.monthly_trend():
        ws3.append([r["month"], round((r["income"] or 0) / 100, 2),
                    round((r["expense"] or 0) / 100, 2),
                    round(((r["income"] or 0) - (r["expense"] or 0)) / 100, 2)])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
